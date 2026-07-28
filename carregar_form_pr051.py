"""
Carrega conferências REAIS da aba FORM-PR-051 da planilha RO-PR-051 para a
RDR_APP_CONFERENCIA. Diferente do seed_conferencias.py (dados fabricados),
aqui os dados são os históricos digitados no formulário.

Percorre TODAS as planilhas RO-PR-051*.xlsx da pasta, em ordem alfabética.
A deduplicação é por (pedido, item, nota_fiscal), tanto contra o que já está
no banco quanto entre as próprias planilhas — rodar de novo não duplica.

Uso — SEMPRE no terminal, nunca pelo Code Runner:

    python carregar_form_pr051.py                        -> simula tudo
    python carregar_form_pr051.py --quantos 200          -> simula 200
    python carregar_form_pr051.py --aplicar              -> grava tudo
    python carregar_form_pr051.py --arquivo "nome.xlsx"  -> só um arquivo

Mapa das colunas (linha 6 é o cabeçalho; dados começam na 7):
  A data inspeção   E nota fiscal    I plano controle   M..V os 10 checks
  D quantidade      F pedido-item    J material         W aprovadas
  B fornecedor(view)G corrida        K EE (spec)        X reprovadas
  C descrição(view) H certificado    L instrumento(s)   Y RPNC   Z inspetor

Instrumentos: os códigos vêm sujos (IM-0023 vs IM0023). Normalizamos
removendo tudo que não é letra/dígito e MAIÚSCULA (IM-0023 -> IM0023),
criando na RDR_APP_LISTA_INSTRUMENTOS o que faltar.

O login runtime (Relatorio) tem INSERT mas NÃO tem DELETE — para desfazer
é preciso USERBI/DBA (filtrar por responsavel ou pelas datas de inspeção).
"""

import datetime
import glob
import os
import re
import sys

import openpyxl

from app import app, db
from app.models import (Recebimento, Conferencia, Corrida, InstrumentoMedicao,
                        Material, PlanoControle, ListaInstrumentos)
from app.metodos import processo_sugerido

APLICAR = "--aplicar" in sys.argv
QUANTOS = None
if "--quantos" in sys.argv:
    QUANTOS = int(sys.argv[sys.argv.index("--quantos") + 1])

if "--arquivo" in sys.argv:
    PLANILHAS = [sys.argv[sys.argv.index("--arquivo") + 1]]
else:
    PLANILHAS = sorted(glob.glob('RO-PR-051*.xlsx'))

ABA = 'FORM-PR-051'

# índices 0-based das colunas
DATA, QTD, NF, PEDIDO, CORRIDA, CERT, PLANO, MAT, EE, INSTR = 0, 3, 4, 5, 6, 7, 8, 9, 10, 11
APROV, REPROV, RPNC, INSPETOR = 22, 23, 24, 25
# os 10 checks, em ordem, mapeados para os atributos do modelo
CHECKS = [
    (12, 'analise_certificado'), (13, 'analise_visual'),
    (14, 'identif_e_rastreabilidade'), (15, 'dimensional'),
    (16, 'dureza_sha'), (17, 'id_ligas'), (18, 'dureza_tt'),
    (19, 'ranhura'), (20, 'fios18_21'), (21, 'rugosidade_acabamento'),
]


def s(v):
    return str(v).strip() if v is not None else ''


def limpo(v):
    """String ou None para placeholders vazios ('-', 'N/A')."""
    t = s(v)
    return None if t in ('', '-', 'N/A', 'n/a') else t


def para_int(v, default=0):
    t = s(v)
    if t in ('', '-', 'N/A'):
        return default
    try:
        return int(float(t))
    except ValueError:
        return default


def para_bool(v):
    t = s(v)
    if t == 'True':
        return True
    if t == 'False':
        return False
    return None   # célula vazia / 'N/A' -> não se aplica


def normalizar_instrumento(codigo):
    """IM-0023 -> IM0023 ; remove tudo que não é letra/dígito, MAIÚSCULA."""
    return re.sub(r'[^A-Za-z0-9]', '', codigo).upper()


with app.app_context():

    materiais_nome = {m.nome.strip(): m for m in db.session.scalars(db.select(Material))}
    materiais_ee = {m.especificacao.strip(): m
                    for m in db.session.scalars(db.select(Material))}
    planos = {p.nome.strip(): p for p in db.session.scalars(db.select(PlanoControle))}
    instrumentos = {i.nome.strip(): i
                    for i in db.session.scalars(db.select(ListaInstrumentos))}

    existentes = set(db.session.execute(db.select(
        Conferencia.pedido, Conferencia.item, Conferencia.nota_fiscal)).all())

    # Mapa da view em memória: (pedido, item) -> {nf_numerica: (nf_da_view, produto)}
    # Serve para dois problemas de uma vez:
    #  1. a NF da planilha às vezes tem um zero a mais ('0000006564' com 10
    #     caracteres não cabe no varchar(9)) — casando pelo VALOR numérico a
    #     grafia canônica é a da view;
    #  2. evita um SELECT por linha para descobrir o produto (2.5k queries).
    mapa_view = {}
    for p, i, nf, produto in db.session.execute(db.select(
            Recebimento.pedido, Recebimento.item,
            Recebimento.nota_fiscal, Recebimento.produto)):
        try:
            mapa_view.setdefault((p, i), {})[int(nf)] = (nf, produto)
        except (TypeError, ValueError):
            continue

    # processo_sugerido() consulta RDR_PROCESSO a cada chamada; como o produto
    # se repete muito entre as linhas, memoriza o resultado por produto.
    cache_processo = {}

    def processo_do_produto(produto):
        if produto not in cache_processo:
            cache_processo[produto] = processo_sugerido(produto)
        return cache_processo[produto]

    criados = 0
    instr_criados = []
    sem_material = 0
    pulados_dup = 0
    orfaos = []
    nf_corrigidas = []
    vistos = set()
    por_planilha = []

    for caminho in PLANILHAS:
        if criados == QUANTOS:
            break

        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
        if ABA not in wb.sheetnames:
            print(f"  {os.path.basename(caminho)}: sem a aba {ABA}, pulada")
            wb.close()
            continue
        ws = wb[ABA]
        linhas = [r for i, r in enumerate(ws.iter_rows(values_only=True), 1)
                  if i >= 7 and isinstance(r[DATA], datetime.datetime)]
        wb.close()

        criados_antes = criados

        for r in linhas:
            if criados == QUANTOS:
                break

            m = re.match(r'^(\d+)\D+(\d+)$', s(r[PEDIDO]))
            if not m:
                continue
            pedido = m.group(1).zfill(6)
            item = m.group(2).zfill(4)
            nf_planilha = s(r[NF])

            # Resolve a NF contra a view. Sem correspondência, a conferência
            # seria órfã: como homepage e gráficos usam INNER JOIN com a view,
            # ela nunca apareceria em tela nenhuma. Melhor não inserir — o
            # login runtime não tem DELETE para desfazer.
            nfs_do_pedido = mapa_view.get((pedido, item), {})
            try:
                nf, produto = nfs_do_pedido[int(nf_planilha)]
            except (KeyError, TypeError, ValueError):
                orfaos.append((pedido, item, nf_planilha,
                               sorted(v[0] for v in nfs_do_pedido.values())))
                continue

            if nf != nf_planilha:
                nf_corrigidas.append((pedido, item, nf_planilha, nf))

            chave = (pedido, item, nf)

            if chave in existentes or chave in vistos:
                pulados_dup += 1
                continue
            vistos.add(chave)

            qt = para_int(r[QTD])

            # material: por nome, depois por código EE
            material = (materiais_nome.get(s(r[MAT]))
                        or materiais_ee.get(s(r[EE])))
            if material is None:
                sem_material += 1

            plano = planos.get(s(r[PLANO]))

            # processo: derivado do produto REAL da view (o form não tem)
            pid = processo_do_produto(produto)

            conf = Conferencia(
                pedido=pedido, item=item, nota_fiscal=nf,
                dt_hr_inspecao=r[DATA],
                qt_total=qt,
                certificado=(limpo(r[CERT]) or '')[:30] or None,
                plano_controle_id=plano.id if plano else None,
                processo_id=pid or None,
                material_id=material.id if material else None,
                pecas_aprovadas=para_int(r[APROV]),
                pecas_reprovadas=para_int(r[REPROV]),
                rpnc=(limpo(r[RPNC]) or '')[:10] or None,
                responsavel=(s(r[INSPETOR]) or 'IMPORTADO')[:50],
            )
            for idx, atributo in CHECKS:
                setattr(conf, atributo, para_bool(r[idx]))

            # corrida (uma, com a quantidade total)
            corrida_txt = limpo(r[CORRIDA])
            if corrida_txt:
                conf.corridas.append(
                    Corrida(corrida=corrida_txt[:50], qtd_corrida=qt))

            # instrumentos: normaliza, cria o que faltar, vincula (sem repetir)
            ids_vinculados = set()
            for bruto in re.findall(r'[A-Za-z]{2}\s*-?\s*\d{2,5}', s(r[INSTR])):
                cod = normalizar_instrumento(bruto)
                if not cod:
                    continue
                inst = instrumentos.get(cod)
                if inst is None:
                    inst = ListaInstrumentos(nome=cod, descricao=None)
                    db.session.add(inst)
                    db.session.flush()          # garante o id
                    instrumentos[cod] = inst
                    instr_criados.append(cod)
                if inst.id not in ids_vinculados:
                    conf.instrumentos.append(
                        InstrumentoMedicao(instrumento_id=inst.id))
                    ids_vinculados.add(inst.id)

            db.session.add(conf)
            criados += 1

        por_planilha.append((os.path.basename(caminho), len(linhas),
                             criados - criados_antes))

    print("\nPor planilha (linhas válidas -> novas a inserir):")
    for nome, lidas, novas in por_planilha:
        curto = nome.replace('RO-PR-051_Relatório_Diário_de_Recebimento ', '')
        print(f"  {curto:<28} {lidas:>5} -> {novas:>5}")

    print(f"\nConferências a inserir: {criados}")
    print(f"  sem material resolvido (material_id NULL): {sem_material}")
    print(f"  linhas puladas por chave duplicada: {pulados_dup}")
    print(f"  NFs corrigidas pela grafia da view: {len(nf_corrigidas)}")
    for pedido, item, antes, depois in nf_corrigidas[:20]:
        print(f"      {pedido}/{item}  {antes} -> {depois}")
    if len(nf_corrigidas) > 20:
        print(f"      ... e mais {len(nf_corrigidas) - 20}")
    print(f"  instrumentos NOVOS criados: "
          f"{sorted(set(instr_criados))} ({len(set(instr_criados))})")

    print(f"\nPULADAS por NF inexistente na view: {len(orfaos)}")
    print(f"  {'PEDIDO/ITEM':<14} {'NF NA PLANILHA':<16} NFs QUE A VIEW TEM")
    for pedido, item, nf_planilha, disponiveis in sorted(orfaos)[:30]:
        print(f"  {pedido}/{item:<7} {nf_planilha:<16} "
              f"{', '.join(disponiveis) if disponiveis else '(pedido/item não existe na view)'}")
    if len(orfaos) > 30:
        print(f"  ... e mais {len(orfaos) - 30} linhas órfãs")

    if APLICAR:
        db.session.commit()
        print("\nGravado.")
    else:
        db.session.rollback()
        print("\nSimulação: nada gravado. Rode com --aplicar para valer.")
